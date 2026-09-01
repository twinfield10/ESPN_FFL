"""Import The Athletic's season projection workbook into a tidy stat table.

**A hand-dropped file, not a scrape.** The workbook is a paid download from The
Athletic (Jake Ciely's spreadsheet) with no API behind it, so this runs when a new
copy is saved rather than nightly. Everything else about it matches the other
season sources: raw stat lines in, ``Data/Projections/TheAthletic/Season/<season>/``
out, and :func:`Scripts.season_projections.load_theathletic_season` reads it from
there.

Why the *team* tabs rather than the flattened ``QB``/``RB``/``WR``/``TE`` ones: the
team tabs are the model. Each is a team-budget times usage-share calculation --
``PASS ATT = team_pass_attempts * player_pass_share`` -- and the per-position tabs
are ``VLOOKUP``s off them that silently drop columns. The QB tab has no receiving
columns at all, which is what hides the one real bug in the file (see
:data:`POSITION_STATS`).

What this deliberately does not read:

- ``FPS``, ``Custom``, ``VORP``, ``AUC$`` -- points and derived values. We score raw
  stats through each league's own rules like every other source, and two of those
  four are wrong anyway: ``OVR & VORP Ranks`` adds 45% of the *row-aligned* running
  back's VORP to each quarterback's, and the QB replacement rank resolves to 2 in a
  one-QB league.
- The ``DST`` tab. ``Settings`` defines all seven points-allowed tiers but the
  ``0 PT GAMES``..``35+ PT GAMES`` columns are null for all 32 teams, so the
  workbook's own defence values omit the points-allowed component entirely. This
  repo's ``DST`` model is blended at 0.25 and is the better number.
- ``Jake's Ranks`` -- a human overlay that deliberately disagrees with his own
  projections. A rank is not a stat line and has nowhere to go in the blend.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Dict, List, Optional, Sequence

import pandas as pd

from Scripts.paths import landing_dir, season_dir

#: Provider directory name under ``Data/Projections``.
SOURCE = "TheAthletic"

#: Blend prefix. Uppercase, because ``compute_weighted_stats`` and ``proj_to_score``
#: scan every uppercase prefix and require it to be numeric.
PREFIX = "ATH"

#: Output filename, matching the FantasyPros season file's shape.
FILENAME = f"{SOURCE}_Projections_Season.parquet"

#: The 32 team tabs, in workbook order. Already on ESPN's abbreviations -- ``WSH``,
#: ``JAX``, ``LV``, ``LAR``, ``LAC`` -- so no alias map is needed here, unlike the
#: nflverse-keyed sources that need :data:`Scripts.draft.board.ESPN_TEAM_ALIASES`.
TEAM_TABS: Sequence[str] = (
    "ARI", "ATL", "BAL", "BUF", "CAR", "CHI", "CIN", "CLE",
    "DAL", "DEN", "DET", "GB", "HOU", "IND", "JAX", "KC",
    "LV", "LAC", "LAR", "MIA", "MIN", "NE", "NO", "NYG",
    "NYJ", "PHI", "PIT", "SF", "SEA", "TB", "TEN", "WSH",
)

#: Workbook column header -> ESPN stat name.
#:
#: The same twelve stats FantasyPros supplies, under ESPN's names, so both sources
#: land on identical ``<PREFIX>_<stat>`` columns and the blend compares like with
#: like. See ``Scripts/scrape_FP.py`` ``final_cols`` for the other end of it.
#:
#: **No ``lostFumbles``.** The workbook does not project fumbles, so ``ATH_`` simply
#: has no column for it and ``compute_weighted_stats`` renormalises the sources that
#: do -- the same handling a book with no line on a stat already gets.
STAT_COLUMNS: Dict[str, str] = {
    "PASS ATT": "passingAttempts",
    "COMP": "passingCompletions",
    "PASS YARDS": "passingYards",
    "PASS TD": "passingTouchdowns",
    "INT": "passingInterceptions",
    "RUSH ATT": "rushingAttempts",
    "RUSH YARDS": "rushingYards",
    "RUSH TD": "rushingTouchdowns",
    "TARGETS": "receivingTargets",
    "REC": "receivingReceptions",
    "RECV YARDS": "receivingYards",
    "RECV TD": "receivingTouchdowns",
}

#: Which stats each position is allowed to carry, and why this is not paranoia.
#:
#: The workbook allocates team target share across a tab's rows, and on the New
#: Orleans tab some of it lands on a **quarterback**: Spencer Rattler carries 32.2
#: targets, 23.7 receptions, 258.7 receiving yards and 2.38 receiving touchdowns.
#: The workbook's own ``QB`` tab has no receiving columns so it never sees them and
#: scores him 7.8; read the team tab straight and he scores **59.8**, which would
#: make a third-string quarterback a real opinion in the blend.
#:
#: Verified across all 32 tabs as of the 2026-08-31 workbook: exactly one player is
#: affected, and no running back, receiver or tight end carries passing stats. The
#: mask is kept anyway -- it is the share model that produced this, so the next
#: download can produce it somewhere else.
POSITION_STATS: Dict[str, frozenset] = {
    "QB": frozenset({
        "passingAttempts", "passingCompletions", "passingYards",
        "passingTouchdowns", "passingInterceptions",
        "rushingAttempts", "rushingYards", "rushingTouchdowns",
    }),
    "RB": frozenset({
        "rushingAttempts", "rushingYards", "rushingTouchdowns",
        "receivingTargets", "receivingReceptions", "receivingYards",
        "receivingTouchdowns",
    }),
    "WR": frozenset({
        "rushingAttempts", "rushingYards", "rushingTouchdowns",
        "receivingTargets", "receivingReceptions", "receivingYards",
        "receivingTouchdowns",
    }),
    "TE": frozenset({
        "receivingTargets", "receivingReceptions", "receivingYards",
        "receivingTouchdowns",
    }),
}

#: Positions read out of the team tabs. Kickers and defences are not on them.
POSITIONS: Sequence[str] = tuple(POSITION_STATS)


def _header_map(row: Sequence) -> Dict[str, int]:
    """Column index for each header on a team tab.

    Read per tab rather than hard-coded, so a workbook that gains a column shifts
    nothing. Verified identical across all 32 tabs, but reading it is free.

    Args:
        row: The tab's first row, as values.

    Returns:
        Dict[str, int]: Header text -> zero-based column index.
    """
    return {str(cell).strip(): i
            for i, cell in enumerate(row) if cell is not None}


def read_workbook(path: Path) -> pd.DataFrame:
    """Parse the 32 team tabs into one tidy row per player.

    Args:
        path: The ``.xlsx`` workbook.

    Returns:
        pd.DataFrame: ``player_name``, ``pro_team``, ``position``, ``bye`` and one
        column per ESPN stat name, masked to what the position can hold.

    Raises:
        FileNotFoundError: If ``path`` does not exist.
        KeyError: If a team tab is missing or its header lacks ``PLAYER``/``POS``.
    """
    import openpyxl

    if not path.exists():
        raise FileNotFoundError(path)

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows: List[dict] = []
        for tab in TEAM_TABS:
            if tab not in book.sheetnames:
                raise KeyError(f"team tab {tab!r} missing from {path.name}")
            sheet = book[tab]
            header: Optional[Dict[str, int]] = None
            for raw in sheet.iter_rows(values_only=True):
                if header is None:
                    header = _header_map(raw)
                    if "PLAYER" not in header or "POS" not in header:
                        raise KeyError(f"{tab}: header has no PLAYER/POS column")
                    continue
                name = raw[header["PLAYER"]] if header["PLAYER"] < len(raw) else None
                pos = raw[header["POS"]] if header["POS"] < len(raw) else None
                if not name or not pos:
                    # Blank spacer rows separate the position blocks, and the
                    # team-totals block sits below them with no POS at all.
                    continue
                pos = str(pos).strip().upper()
                if pos not in POSITION_STATS:
                    continue
                allowed = POSITION_STATS[pos]
                record = {
                    "player_name": str(name).strip(),
                    "pro_team": tab,
                    "position": pos,
                    "bye": raw[header["BYE"]] if "BYE" in header else None,
                }
                masked = []
                for head, stat in STAT_COLUMNS.items():
                    idx = header.get(head)
                    value = raw[idx] if idx is not None and idx < len(raw) else None
                    if stat in allowed:
                        record[stat] = value
                        continue
                    record[stat] = None
                    if value is not None and float(value or 0.0) != 0.0:
                        masked.append(stat)
                # Recorded rather than merely dropped, so `audit` can name what the
                # mask caught. Measuring the output frame cannot: by then the
                # offending values are already None and every count reads zero.
                record["masked_stats"] = ",".join(masked)
                rows.append(record)
    finally:
        book.close()

    frame = pd.DataFrame(rows)
    for stat in STAT_COLUMNS.values():
        frame[stat] = pd.to_numeric(frame[stat], errors="coerce")
    return frame


#: Columns of :func:`read_workbook` that are diagnostics rather than stats.
#:
#: ``masked_stats`` is lowercase and stays out of the ``ATH_`` namespace on purpose:
#: ``UPPER_`` is reserved for blendable numerics, and both ``compute_weighted_stats``
#: and ``proj_to_score`` scan every uppercase prefix and require it to be numeric.
#: :func:`Scripts.season_projections.load_theathletic_season` reads only the stat
#: columns, so this never reaches the board.
DIAGNOSTIC_COLUMNS: Sequence[str] = ("masked_stats",)


def audit(frame: pd.DataFrame, top: int = 12) -> None:
    """Print what parsed and how much of it the ID crosswalk can confirm.

    The join into the board is by **name** -- ``normalise_name`` then ``join_key``,
    the same path FantasyPros takes -- because only 289 of the workbook's players
    carry an ID at all. This is therefore an audit rather than the join: it says how
    much of the file the crosswalk agrees exists, and names the biggest players it
    cannot place, which is where a name-join failure would cost the most.

    Args:
        frame: Output of :func:`read_workbook`.
        top: How many unmatched names to name.
    """
    counts = frame["position"].value_counts().to_dict()
    print(f"  parsed {len(frame)} players: "
          + " / ".join(f"{p} {counts.get(p, 0)}" for p in POSITIONS))

    bled = frame[frame["masked_stats"].astype(str).str.len() > 0]
    print(f"  position mask: dropped off-position stats from {len(bled)} rows")
    for _, row in bled.iterrows():
        print(f"    masked: {row['player_name']} ({row['position']}, "
              f"{row['pro_team']}) -- {row['masked_stats']}")

    try:
        from Scripts.crosswalk import load_crosswalk
        from Scripts.season_projections import normalise_name
    except Exception as exc:                                   # pragma: no cover
        print(f"  crosswalk audit skipped: {exc}")
        return

    try:
        cross = load_crosswalk()
    except Exception as exc:
        print(f"  crosswalk audit skipped: {exc}")
        return

    known = {normalise_name(n) for n in cross["name"].to_list() if n}
    keys = frame["player_name"].map(normalise_name)
    missing = frame.loc[~keys.isin(known)].copy()
    matched = len(frame) - len(missing)
    print(f"  crosswalk: {matched}/{len(frame)} names resolve "
          f"({100 * matched / max(len(frame), 1):.0f}%)")
    if len(missing):
        order = [c for c in ("receivingYards", "rushingYards", "passingYards")
                 if c in missing.columns]
        missing["_size"] = missing[order].fillna(0).sum(axis=1)
        worst = missing.sort_values("_size", ascending=False).head(top)
        for _, row in worst.iterrows():
            print(f"    unmatched: {row['player_name']} "
                  f"({row['position']}, {row['pro_team']})")


def build(season: int, path: Path) -> pd.DataFrame:
    """Import the workbook and write the season files.

    Keeps the ``.xlsx`` under ``Landing/`` unmodified, so the tidy table can always
    be rebuilt from what was actually downloaded, and writes both parquet and csv
    like every other source -- parquet is authoritative, the csv is for eyeballing.

    Args:
        season: Season year, for the output path.
        path: The workbook to read.

    Returns:
        pd.DataFrame: The tidy table that was written.
    """
    frame = read_workbook(path)
    audit(frame)

    kept = landing_dir(SOURCE, season, path.name)
    if path.resolve() != kept.resolve():
        shutil.copy2(path, kept)
    print(f"  landed {kept.relative_to(kept.parents[4])}")

    out = season_dir(SOURCE, season, FILENAME)
    frame.to_parquet(out)
    frame.to_csv(out.with_suffix(".csv"), index=False)
    print(f"The Athletic season-long {season}: {len(frame)} rows, "
          f"{frame['player_name'].nunique()} players -> {out.name}")
    return frame


def main(argv=None):
    """Command-line entry point.

    No side effects at import: the workbook is only read when this is called, so
    importing the module cannot overwrite a season's file.
    """
    import argparse

    from Scripts.nfl_utils import current_season

    p = argparse.ArgumentParser(
        prog="python -m Scripts.load_athletic",
        description="Import The Athletic's projection workbook (a manual download).",
    )
    p.add_argument("--season", type=int, default=None,
                   help="defaults to the schedule's season")
    p.add_argument("--file", required=True, type=Path,
                   help="path to the .xlsx workbook")
    args = p.parse_args(argv)

    season = current_season() if args.season is None else args.season
    build(int(season), args.file)


if __name__ == "__main__":
    main()
